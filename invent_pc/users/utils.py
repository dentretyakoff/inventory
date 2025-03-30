import openpyxl
from django_filters import FilterSet
from django.db.models import Model

from utils.utils import get_counters, get_pages
from .models import ADUsers, Radius, VPN, StatusChoices


def update_or_create_users(
        model: Model, users_data: list[dict[str, str]]) -> None:
    """Обновляет пользователей или создает новых."""
    existing_users = model.objects.in_bulk(field_name='login')
    new_users = []

    for user_data in users_data:
        login = user_data.pop('login')
        user = existing_users.pop(login, None)
        user_data['successfully_updated'] = True
        if not user:
            new_users.append(model(login=login, **user_data))
        elif user.needs_update(user_data):
            user.save()

    if existing_users:
        for existing_user in existing_users.values():
            existing_user.successfully_updated = False
        model.objects.bulk_update(existing_users.values(),
                                  ['successfully_updated'])

    if new_users:
        model.objects.bulk_create(new_users)


def make_clean_wb(title: str, headers: list[str]) -> openpyxl.Workbook:
    """Формирует чистую книгу Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)

    for cell in ws['1:1']:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='00C0C0C0')

    return wb


def get_file(users: ADUsers) -> openpyxl.Workbook:
    """Создает Excel со связными учетными записями."""
    title = 'Список учетных записей'
    headers = ['№', 'ФИО', 'Логин', 'Email',
               'Wi-Fi', 'VPN логин', 'VPN комментарий']
    wb = make_clean_wb(title, headers)
    ws = wb.active

    for num, user in enumerate(users, 1):
        ws.append([
            num,
            user.fio,
            user.login,
            user.email,
            getattr(user.rdlogin, 'login', '-'),
            getattr(user.vpn, 'login', '-'),
            getattr(user.vpn, 'comment', '-')
        ])

        statuses = {
            # ws.cell(row= , column=)
            ws.cell(num+1, 3): getattr(user, 'status', None),
            ws.cell(num+1, 5): getattr(user.rdlogin, 'status', None),
            ws.cell(num+1, 6): getattr(user.vpn, 'status', None)
        }
        for cell, status in statuses.items():
            if status == StatusChoices.INACTIVE:
                cell.fill = openpyxl.styles.PatternFill(
                    'solid', fgColor='FF0000')  # Красный

    return wb


def get_radius_file(users: Radius) -> openpyxl.Workbook:
    """Создает Excel со свободными учетными записями Radius."""
    title = 'Список учетных записей'
    headers = ['№', 'ФИО', 'Логин']
    wb = make_clean_wb(title, headers)
    ws = wb.active

    for num, user in enumerate(users, 1):
        ws.append([num, user.fio, user.login])
        if user.status == StatusChoices.INACTIVE:
            ws.cell(num+1, 3).fill = openpyxl.styles.PatternFill(
                'solid', fgColor='FF0000')  # Красный

    return wb


def get_vpn_file(users: VPN) -> openpyxl.Workbook:
    """Создает Excel со свободными учетными записями VPN."""
    title = 'Список учетных записей'
    headers = ['№', 'Логин', 'Комментарий']
    wb = make_clean_wb(title, headers)
    ws = wb.active

    for num, user in enumerate(users, 1):
        ws.append([num, user.login, user.comment])
        if user.status == StatusChoices.INACTIVE:
            ws.cell(num+1, 2).fill = openpyxl.styles.PatternFill(
                'solid', fgColor='FF0000')  # Красный

    return wb


def get_users_base_context(request, user_filter: FilterSet) -> dict[any]:
    """Формирует базовый context для страниц учетных записей."""
    ad_users = (ADUsers.objects.all()
                .select_related('rdlogin', 'vpn')
                .prefetch_related('gigro'))
    related_statuses = ad_users.values('rdlogin__status', 'vpn__status')
    unrelated_radius_statuses = (Radius.objects
                                 .filter(ad_user__isnull=True)
                                 .values('status'))
    unrelated_vpn_statuses = (VPN.objects
                              .filter(ad_user__isnull=True)
                              .values('status'))

    ad_users_statuses = get_counters(ad_users.values('status'), 'status')
    radius_users_statuses = get_counters(related_statuses, 'rdlogin__status')
    vpn_users_statuses = get_counters(related_statuses, 'vpn__status')

    unrelated_radius_users_statuses = get_counters(
        unrelated_radius_statuses, 'status')
    unrelated_vpn_users_statuses = get_counters(
        unrelated_vpn_statuses, 'status')

    page_obj = get_pages(request, user_filter.qs)
    current_query_params = request.GET.copy()
    current_query_params.pop('page', None)

    context = {
        'page_obj': page_obj,
        'ad_users_statuses': ad_users_statuses,
        'radius_users_statuses': radius_users_statuses,
        'vpn_users_statuses': vpn_users_statuses,
        'unrelated_radius_users_statuses': unrelated_radius_users_statuses,
        'unrelated_vpn_users_statuses': unrelated_vpn_users_statuses,
        'current_query_params': current_query_params,
    }

    return context
